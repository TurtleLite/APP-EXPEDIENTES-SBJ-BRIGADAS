import datetime
import math
import os
from typing import Dict, List, Optional
from sqlalchemy.orm import Session
from app.models.list_definition import ListDefinition, ListRecord

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.properties import PageSetupProperties

LOGO_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), "assets", "logo_sbj.png")

PRIMARY = "6E7B91"
DARK = "3F4650"
BORDER = "D7DBE1"
ALT_ROW = "F4F6F8"
MUTED = "8A919C"

FILTER_LABELS = {
    "especialidad": "Especialidad",
    "perfil": "Perfil",
    "criticidad": "Criticidad Clínica",
    "estatus_cirugia": "Estatus de cirugía",
}


def _format_filters(filt: Optional[dict]) -> str:
    if not filt:
        return "Sin filtros"
    parts = []
    for key, label in FILTER_LABELS.items():
        val = filt.get(key)
        if val not in (None, ""):
            parts.append(f"{label}: {val}")
    return " · ".join(parts) if parts else "Sin filtros"


def _thin_border(color: str = BORDER):
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def _px_to_units(px: int) -> float:
    """Convierte píxeles a unidades de ancho de columna de Excel (1 unidad ≈ 7px)."""
    return round(max(4.0, (px - 5) / 7.0), 2)


KNOWN_WIDTHS = {
    "No": 4.0,
    "Nombre/Name": 32.81,
    "Age": 5.3,
    "Diagnostic/Procedure": 22.0,
    "Pf": 4.43,
    "Origin": 17.86,
    "Phone NO.": 11.57,
    "Housing": 8.1,
    "Chart": 8.72,
    "Referred by": 15.86,
}

def _units_to_px(units: float) -> int:
    return int(round(units * 7 + 5))


def _wrapped_row_height(
    record: Dict,
    columns: List[str],
    widths: List[float],
    line_height: float = 12.75,
) -> float:
    """Calcula la altura que necesita la fila para que el texto envuelto (wrap) quepa
    dentro de la celda, evitando que Excel corte el contenido al imprimir."""
    max_lines = 1
    for col_name, w in zip(columns, widths):
        text = record.get(col_name, None)
        if text is None:
            continue
        text = str(text)
        if not text:
            continue
        capacity = max(1, int(w))
        lines = 0
        for segment in text.split("\n"):
            lines += max(1, math.ceil(len(segment) / capacity))
        max_lines = max(max_lines, lines)
    return round(max_lines * line_height, 2)


def _data_cell_alignment(col_name: str):
    # Si el texto no cabe en la celda, se escribe en otra línea dentro de la celda (wrap text).
    return Alignment(horizontal="center", vertical="center", wrap_text=True, shrink_to_fit=False)


def _apply_fixed_widths(ws, columns: List[str], widths: List[float], data_col0: int):
    """Aplica anchos fijos: Origin 120px y Nombre/Name 180px. Los píxeles liberados
    respecto a los valores previos se suman a Observación (base 100px)."""
    obs_px = 100
    if "Origin" in columns:
        obs_px += max(0, 150 - 120)
        wid = _px_to_units(120)
        widths[columns.index("Origin")] = wid
        ws.column_dimensions[get_column_letter(columns.index("Origin") + data_col0)].width = wid
    if "Nombre/Name" in columns:
        old = _units_to_px(widths[columns.index("Nombre/Name")])
        obs_px += max(0, old - 180)
        wid = _px_to_units(180)
        widths[columns.index("Nombre/Name")] = wid
        ws.column_dimensions[get_column_letter(columns.index("Nombre/Name") + data_col0)].width = wid
    if "Observación" in columns:
        wid = _px_to_units(obs_px)
        widths[columns.index("Observación")] = wid
        ws.column_dimensions[get_column_letter(columns.index("Observación") + data_col0)].width = wid


def _content_widths(records: List[dict], columns: List[str]) -> List[int]:
    widths = []
    for col in columns:
        if col in KNOWN_WIDTHS:
            widths.append(KNOWN_WIDTHS[col])
            continue
        lens = [len(str(rec.get(col, ""))) for rec in records] or [0]
        header = min(len(str(col)), 8)
        longest = max(lens)
        eff = max(header, longest)
        widths.append(min(26, max(4, eff + 2)))
    return widths


def import_records_from_excel(db: Session, list_id: int, filepath: str) -> int:
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    ld = db.query(ListDefinition).filter(ListDefinition.id == list_id).first()
    if not ld:
        raise ValueError("Lista no encontrada")

    headers = [cell.value for cell in ws[1]]
    col_keys = [c["key"] for c in ld.columns_config]
    col_map = {}
    for i, header in enumerate(headers):
        for key in col_keys:
            if header and header.lower() == key.lower():
                col_map[i] = key
                break

    count = 0
    from app.services.record_service import _is_expediente_list, numero_expediente_final
    is_expediente = _is_expediente_list(db, list_id)
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        data = {}
        for i, cell in enumerate(row):
            if cell.value is not None and i in col_map:
                data[col_map[i]] = cell.value
        if data:
            if is_expediente:
                numero = str(data.get("expediente", "") or "")
                import re as _re
                numero = _re.sub(r"\D", "", numero)
                if not numero:
                    rollback_msg = f"Falta el número de expediente en la fila {row_idx} del archivo Excel"
                    raise ValueError(rollback_msg)
                data["expediente"] = numero_expediente_final(db, numero)
            if is_expediente and not data.get("estatus_cirugia"):
                data["estatus_cirugia"] = "En espera"
            record = ListRecord(list_definition_id=list_id, data=data)
            db.add(record)
            count += 1

    db.commit()
    return count


def _now_honduras() -> datetime.datetime:
    return datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=-6))).replace(tzinfo=None)


def _write_header_block(
    ws,
    title: Optional[str],
    count: int,
    last_col: str,
    data_col0: int,
    now: datetime.datetime,
    institution: str = "Centro Médico San Benito José",
    filters: Optional[dict] = None,
) -> int:
    row = 1
    if title:
        ws.column_dimensions["A"].width = 3

        ws.merge_cells(f"B{row}:{last_col}{row}")
        c = ws.cell(row=row, column=2, value=institution.upper())
        c.font = Font(bold=True, size=10, color=DARK)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 34
        row += 1

        ws.merge_cells(f"B{row}:{last_col}{row}")
        c = ws.cell(row=row, column=2, value=title)
        c.font = Font(bold=True, size=15, color=PRIMARY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 34
        row += 1

        if os.path.exists(LOGO_PATH):
            from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
            from openpyxl.drawing.xdr import XDRPositiveSize2D
            from openpyxl.utils.units import pixels_to_EMU
            logo = XLImage(LOGO_PATH)
            ratio = logo.height / logo.width
            logo.width = 120
            logo.height = round(logo.width * ratio)
            header_px = (ws.row_dimensions[1].height + ws.row_dimensions[2].height) * 96 / 72
            row_off = max(0, round((header_px - logo.height) / 2))
            logo.anchor = OneCellAnchor(
                _from=AnchorMarker(col=0, colOff=pixels_to_EMU(30), row=0, rowOff=pixels_to_EMU(row_off)),
                ext=XDRPositiveSize2D(cx=pixels_to_EMU(logo.width), cy=pixels_to_EMU(logo.height)),
            )
            ws.add_image(logo)

        ws.merge_cells(f"B{row}:{last_col}{row}")
        c = ws.cell(
            row=row,
            column=2,
            value=f"Generado el {now.strftime('%d/%m/%Y')} a las {now.strftime('%H:%M')}   |   Total de registros: {count}",
        )
        c.font = Font(size=8, italic=True, color=MUTED)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 20
        row += 1

        ws.row_dimensions[row].height = 2
        row += 1  # spacer

        ws.merge_cells(f"B{row}:{last_col}{row}")
        line = ws.cell(row=row, column=2)
        line.fill = PatternFill(start_color=BORDER, end_color=BORDER, fill_type="solid")
        ws.row_dimensions[row].height = 1.5
        row += 1

        ws.row_dimensions[row].height = 3
        row += 1  # spacer

        filter_text = _format_filters(filters)
        if filter_text != "Sin filtros":
            ws.merge_cells(f"B{row}:{last_col}{row}")
            c = ws.cell(row=row, column=2)
            from openpyxl.cell.rich_text import CellRichText, TextBlock
            from openpyxl.cell.text import InlineFont
            c.value = CellRichText(
                TextBlock(InlineFont(b=True, sz=8, color=DARK), f"Filtros aplicados:  "),
                TextBlock(InlineFont(sz=8, color=MUTED), filter_text),
            )
            c.fill = PatternFill(start_color=ALT_ROW, end_color=ALT_ROW, fill_type="solid")
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = _thin_border()
            ws.row_dimensions[row].height = 13
            row += 1
            ws.row_dimensions[row].height = 4
            row += 1  # spacer

    return row


def export_to_excel(
    records: List[dict],
    columns: List[str],
    filepath: str,
    title: Optional[str] = None,
    filters: Optional[dict] = None,
    count: Optional[int] = None,
    institution: str = "Centro Médico San Benito José",
):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte"

    ncols = len(columns)
    data_col0 = 2 if title else 1
    last_col = get_column_letter(ncols + (1 if title else 0))
    now = _now_honduras()

    ws.sheet_view.showGridLines = False

    row = _write_header_block(
        ws, title, count if count is not None else len(records), last_col, data_col0, now,
        institution=institution, filters=filters,
    )

    header_row = row

    widths = _content_widths(records, columns)
    total_cols = len(widths) + (1 if title else 0)
    col_a_units = 3 if title else 0
    target_units = ((11.0 - 0.8) * 96 - 3 * total_cols) / 7.0
    factor = (target_units - col_a_units) / (sum(widths) or 1)
    widths = [round(w * factor, 2) for w in widths]
    if title:
        missing = target_units - col_a_units - sum(widths)
        if missing > 0.5 and widths:
            base = sum(widths)
            widths = [round(w + missing * w / base, 2) for w in widths]
    for i, w in enumerate(widths, data_col0):
        ws.column_dimensions[get_column_letter(i)].width = w

    _apply_fixed_widths(ws, columns, widths, data_col0)

    for col_idx, col_name in enumerate(columns, 1):
        col = col_idx + (1 if title else 0)
        cell = ws.cell(row=header_row, column=col, value=col_name)
        cell.fill = PatternFill(start_color=PRIMARY, end_color=PRIMARY, fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        side_bottom = Side(style="medium", color="3F4650")
        side_thin = Side(style="thin", color=BORDER)
        cell.border = Border(left=side_thin, right=side_thin, top=side_thin, bottom=side_bottom)
    ws.row_dimensions[header_row].height = 17

    for data_idx, record in enumerate(records, header_row + 1):
        for col_idx, col_name in enumerate(columns, 1):
            col = col_idx + (1 if title else 0)
            val = record.get(col_name, "")
            cell = ws.cell(row=data_idx, column=col, value=val)
            cell.border = _thin_border()
            cell.font = Font(size=10, color="444B54")
            cell.alignment = _data_cell_alignment(col_name)
            if (data_idx - header_row) % 2 == 0:
                cell.fill = PatternFill(start_color=ALT_ROW, end_color=ALT_ROW, fill_type="solid")
        ws.row_dimensions[data_idx].height = _wrapped_row_height(record, columns, widths)

    first_col = get_column_letter(data_col0)
    ws.auto_filter.ref = f"{first_col}{header_row}:{last_col}{header_row + len(records)}"
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_options.horizontalCentered = True
    ws.page_setup.paperSize = 1  # Carta (Letter) 8.5" x 11"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.7
    ws.page_margins.bottom = 0.7

    ws.print_title_rows = f"{header_row}:{header_row}"

    ws.oddFooter.left.text = f"Generado el {now.strftime('%d/%m/%Y')}"
    ws.oddFooter.left.size = 8
    ws.oddFooter.left.color = MUTED
    ws.oddFooter.right.text = "Página &P de &N"
    ws.oddFooter.right.size = 8
    ws.oddFooter.right.color = MUTED

    wb.save(filepath)


def _write_section_table(
    ws,
    row: int,
    esp: str,
    records: List[dict],
    columns: List[str],
    widths: List[float],
    data_col0: int,
    last_col: str,
) -> int:
    first_col = get_column_letter(data_col0)
    ws.merge_cells(f"{first_col}{row}:{last_col}{row}")
    c = ws.cell(row=row, column=data_col0, value=f"{esp}  ·  {len(records)} pacientes")
    c.font = Font(bold=True, size=11, color="FFFFFF")
    c.fill = PatternFill(start_color=DARK, end_color=DARK, fill_type="solid")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 18
    row += 1

    header_row = row
    for col_idx, col_name in enumerate(columns, 1):
        col = col_idx + data_col0 - 1
        cell = ws.cell(row=row, column=col, value=col_name)
        cell.fill = PatternFill(start_color=PRIMARY, end_color=PRIMARY, fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        side_bottom = Side(style="medium", color="3F4650")
        side_thin = Side(style="thin", color=BORDER)
        cell.border = Border(left=side_thin, right=side_thin, top=side_thin, bottom=side_bottom)
    ws.row_dimensions[header_row].height = 17
    row += 1

    for data_idx, record in enumerate(records, row):
        for col_idx, col_name in enumerate(columns, 1):
            col = col_idx + data_col0 - 1
            val = record.get(col_name, "")
            cell = ws.cell(row=data_idx, column=col, value=val)
            cell.border = _thin_border()
            cell.font = Font(size=10, color="444B54")
            cell.alignment = _data_cell_alignment(col_name)
            if (data_idx - header_row) % 2 == 0:
                cell.fill = PatternFill(start_color=ALT_ROW, end_color=ALT_ROW, fill_type="solid")
        ws.row_dimensions[data_idx].height = _wrapped_row_height(record, columns, widths)
    row = data_idx + 1

    ws.row_dimensions[row].height = 6
    return row + 1


def export_to_excel_stream(
    rows,
    columns: List[str],
    filepath: str,
    title: Optional[str] = None,
    count: Optional[int] = None,
    institution: str = "Centro Médico San Benito José",
):
    """Exportación en streaming (memoria constante) para listas muy grandes.
    No incluye logo ni celdas combinadas (limitaciones de openpyxl write_only)."""
    from openpyxl.cell import WriteOnlyCell

    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet("Reporte")
    now = _now_honduras()

    widths = [KNOWN_WIDTHS.get(col, 18.0) for col in columns]
    for i, w in enumerate(widths, 2):
        ws.column_dimensions[get_column_letter(i)].width = w

    def styled(value, font=None, fill=None, align=None):
        cell = WriteOnlyCell(ws, value=value)
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if align:
            cell.alignment = align
        return cell

    if title:
        ws.append([styled(institution.upper(), font=Font(bold=True, size=10, color=DARK))])
        ws.append([styled(title, font=Font(bold=True, size=15, color=PRIMARY))])
        ws.append([styled(
            f"Generado el {now.strftime('%d/%m/%Y')} a las {now.strftime('%H:%M')}   |   Total de registros: {count if count is not None else ''}",
            font=Font(size=8, italic=True, color=MUTED),
        )])
        ws.append([])

    header_cells = []
    for col in columns:
        cell = styled(col, font=Font(color="FFFFFF", bold=True, size=10),
                      fill=PatternFill(start_color=PRIMARY, end_color=PRIMARY, fill_type="solid"),
                      align=Alignment(horizontal="center", vertical="center"))
        header_cells.append(cell)
    ws.append(header_cells)

    for record in rows:
        ws.append([
            styled(
                record.get(col, ""),
                font=Font(size=10, color="444B54"),
                align=_data_cell_alignment(col),
            )
            for col in columns
        ])

    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_options.horizontalCentered = True
    ws.page_setup.paperSize = 1
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.7
    ws.page_margins.bottom = 0.7

    ws.oddFooter.left.text = f"Generado el {now.strftime('%d/%m/%Y')}"
    ws.oddFooter.left.size = 8
    ws.oddFooter.left.color = MUTED
    ws.oddFooter.right.text = "Página &P de &N"
    ws.oddFooter.right.size = 8
    ws.oddFooter.right.color = MUTED

    wb.save(filepath)


def export_day_list_to_excel(
    sections: List[dict],
    columns: List[str],
    filepath: str,
    title: str,
    count: Optional[int] = None,
    institution: str = "Centro Médico San Benito José",
):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Listado"

    ncols = len(columns)
    data_col0 = 2
    last_col = get_column_letter(ncols + 1)
    now = _now_honduras()
    all_rows = [r for sec in sections for r in sec["rows"]]
    count = count if count is not None else len(all_rows)

    ws.sheet_view.showGridLines = False

    row = _write_header_block(ws, title, count, last_col, data_col0, now, institution=institution)

    widths = _content_widths(all_rows, columns)
    total_cols = len(widths) + 1
    target_units = ((11.0 - 0.8) * 96 - 3 * total_cols) / 7.0
    factor = (target_units - 3) / (sum(widths) or 1)
    widths = [round(w * factor, 2) for w in widths]
    missing = target_units - 3 - sum(widths)
    if missing > 0.5 and widths:
        base = sum(widths)
        widths = [round(w + missing * w / base, 2) for w in widths]
    for i, w in enumerate(widths, data_col0):
        ws.column_dimensions[get_column_letter(i)].width = w

    _apply_fixed_widths(ws, columns, widths, data_col0)

    for sec in sections:
        row = _write_section_table(ws, row, sec["esp"], sec["rows"], columns, widths, data_col0, last_col)

    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_options.horizontalCentered = True
    ws.page_setup.paperSize = 1  # Carta (Letter) 8.5" x 11"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.7
    ws.page_margins.bottom = 0.7

    ws.oddFooter.left.text = f"Generado el {now.strftime('%d/%m/%Y')}"
    ws.oddFooter.left.size = 8
    ws.oddFooter.left.color = MUTED
    ws.oddFooter.right.text = "Página &P de &N"
    ws.oddFooter.right.size = 8
    ws.oddFooter.right.color = MUTED

    wb.save(filepath)
