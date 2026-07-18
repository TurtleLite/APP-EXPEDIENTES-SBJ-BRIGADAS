from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy.orm import Session
from app.models.list_definition import ListDefinition, ListRecord
from app.schemas.list_definition import ListDefinitionCreate



EXPEDIENTE_COLUMNS = [
    {"key": "centro_medico", "label": "Centro Médico", "type": "text"},
    {"key": "especialidad", "label": "Especialidad", "type": "text"},
    {"key": "criticidad", "label": "Criticidad Clínica", "type": "text"},
    {"key": "estatus", "label": "Estatus del Paciente", "type": "text"},
    {"key": "nombre", "label": "Nombre/First Name", "type": "text"},
    {"key": "apellido", "label": "Apellido/Last Name", "type": "text"},
    {"key": "sexo", "label": "Sexo/Sex", "type": "text"},
    {"key": "edad", "label": "Age/Edad", "type": "number"},
    {"key": "fecha_elaboracion", "label": "Fecha de Elaboración", "type": "date"},
    {"key": "identidad", "label": "Nº Identidad", "type": "text"},
    {"key": "persona_responsable", "label": "Persona Responsable", "type": "text"},
    {"key": "procedencia", "label": "Procedencia", "type": "text"},
    {"key": "albergue", "label": "Albergue", "type": "text"},
    {"key": "perfil", "label": "Perfil", "type": "text"},
    {"key": "telefono", "label": "Teléfono", "type": "text"},
    {"key": "expediente", "label": "Expediente", "type": "text"},
    {"key": "domicilio", "label": "Domicilio del Paciente", "type": "text"},
    {"key": "historia_enfermedad", "label": "Historia de Enfermedad Actual", "type": "text"},
    {"key": "enfermedades_previas", "label": "Enfermedades Anteriores", "type": "text"},
    {"key": "cirugias_previas", "label": "Cirugías Anteriores", "type": "text"},
    {"key": "alergias", "label": "Alergias", "type": "text"},
    {"key": "otros_antecedentes", "label": "Otros Antecedentes", "type": "text"},
    {"key": "presion_arterial", "label": "P.A./B.P", "type": "text"},
    {"key": "fc", "label": "F.C.", "type": "text"},
    {"key": "pulso", "label": "Pulso", "type": "text"},
    {"key": "temperatura", "label": "T°", "type": "text"},
    {"key": "fr", "label": "F.R.", "type": "text"},
    {"key": "peso", "label": "Peso/Weight", "type": "text"},
    {"key": "talla", "label": "Talla", "type": "text"},
    {"key": "bmi", "label": "B.M.I.", "type": "text"},
    {"key": "examen_fisico", "label": "Examen Físico", "type": "text"},
    {"key": "diagnostico", "label": "Diagnóstico", "type": "text"},
    {"key": "nombre_medico", "label": "Nombre del Médico", "type": "text"},
    {"key": "cirujano", "label": "Cirujano", "type": "text"},
    {"key": "fecha_cirugia", "label": "Fecha de Cirugía", "type": "date"},
]


def create_expediente_template(db: Session, user_id: int) -> ListDefinition:
    from app.services.list_service import create_list_definition
    schema = ListDefinitionCreate(
        name="Expediente Médico",
        description="Historial clínico de pacientes con datos personales, antecedentes, signos vitales y diagnóstico",
        columns_config=EXPEDIENTE_COLUMNS,
    )
    return create_list_definition(db, schema, user_id)


def _thin_border():
    return Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

def _medium_bottom():
    return Border(bottom=Side(style='medium'))

def _medium_top():
    return Border(top=Side(style='medium'))

def _thin_border_right_medium():
    return Border(left=Side(style='thin'), right=Side(style='medium'), top=Side(style='thin'), bottom=Side(style='thin'))

def _thin_border_left_medium():
    return Border(left=Side(style='medium'), top=Side(style='thin'), bottom=Side(style='thin'))

def _thin_border_bottom_only():
    return Border(bottom=Side(style='thin'))

def _thin_border_top_bottom():
    return Border(top=Side(style='thin'), bottom=Side(style='thin'))

def _thin_border_top():
    return Border(top=Side(style='thin'))

def _thin_border_bottom():
    return Border(bottom=Side(style='thin'))

def _thin_border_sides():
    return Border(left=Side(style='thin'), right=Side(style='thin'))

def _thin_border_left():
    return Border(left=Side(style='thin'))

def _thin_border_right():
    return Border(right=Side(style='thin'))


def export_expediente_excel(records: list[ListRecord], filepath: str, logo_path: str = None):
    import os
    from openpyxl.drawing.image import Image

    wb = Workbook()
    ws = wb.active
    ws.title = "Expediente"
    ws.page_setup.orientation = 'portrait'

    # Styles
    arial = 'Arial'
    center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
    center_nowrap = Alignment(horizontal="center", vertical="center")
    center_vwrap = Alignment(vertical="center", wrap_text=True)
    left_center_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right_center_wrap = Alignment(horizontal="right", vertical="center", wrap_text=True)

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    thin = _thin_border()
    thin_top_bottom = _thin_border_top_bottom()
    thin_top = _thin_border_top()
    thin_bottom = _thin_border_bottom()
    thin_sides = _thin_border_sides()
    thin_left = _thin_border_left()
    thin_right = _thin_border_right()
    thin_right_medium = _thin_border_right_medium()
    thin_left_medium = _thin_border_left_medium()
    medium_bottom = _medium_bottom()
    medium_top = _medium_top()

    col_widths = {
        'A': 11.43, 'B': 16.0, 'C': 11.43, 'D': 16.0,
        'E': 9.43, 'F': 10.14, 'G': 11.71, 'H': 14.86,
        'I': 15.29, 'J': 13.71,
    }

    row_heights = {
        1: 20.25, 2: 24.0, 3: 26.25, 5: 12.75, 6: 12.75,
        8: 12.75, 9: 12.75, 10: 12.75, 14: 13.5, 15: 12.75,
        16: 12.75, 21: 12.75, 25: 12.75, 26: 12.75, 27: 12.75,
        29: 12.75, 30: 12.75, 38: 12.75, 44: 18.75, 48: 13.5, 50: 13.5,
    }

    if logo_path is None:
        _try_paths = [
            os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'assets', 'logo_sbj.png'),
            os.path.join(os.getcwd(), 'backend', 'app', 'assets', 'logo_sbj.png'),
            os.path.join(os.getcwd(), 'app', 'assets', 'logo_sbj.png'),
        ]
        logo_path = None
        for p in _try_paths:
            if os.path.exists(p):
                logo_path = p
                break

    for row_idx, record in enumerate(records):
        d = record.data if record.data else {}

        if row_idx > 0:
            ws.append([])
            r = ws.max_row + 1
        else:
            r = 1

        base = r

        # === ROW 1: Title + Logo ===
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        c = ws.cell(r, 1, d.get("centro_medico", "Centro Médico"))
        c.font = Font(name=arial, bold=True, size=16)
        c.alignment = center_wrap
        if os.path.exists(logo_path):
            img = Image(logo_path)
            img.anchor = f'A{r}'
            ws.add_image(img)
        r += 1

        # === ROW 2: Especialidad / Criticidad / Estatus ===
        ws.merge_cells(start_row=r, start_column=7, end_row=r, end_column=8)
        c = ws.cell(r, 7, "Especialidad")
        c.font = Font(name=arial, bold=True, size=16, color="FF0000")
        c.alignment = center_wrap

        c = ws.cell(r, 9, "Criticidad clínica")
        c.font = Font(name=arial, bold=True, size=10)
        c.fill = yellow_fill
        c.alignment = Alignment(vertical="center", wrap_text=True)

        c = ws.cell(r, 10, "Estatus de paciente")
        c.font = Font(name=arial, bold=True, size=10)
        c.fill = yellow_fill
        c.alignment = center_wrap

        ws.cell(r, 1).font = Font(name=arial, bold=True, color="FF0000")
        _apply_borders_range(ws, r, 7, r, 8, thin)
        _apply_border(ws, r, 9, thin)
        _apply_border(ws, r, 10, thin)
        r += 1

        # === ROW 3: Labels ===
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        c = ws.cell(r, 1, "Nombre/First Name")
        c.font = Font(name=arial, bold=True, size=10)
        c.alignment = center_nowrap
        ws.cell(r, 2).border = thin_right_medium

        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        c = ws.cell(r, 3, "Apellido/ Last Name")
        c.font = Font(name=arial, bold=True, size=10)
        c.alignment = center_nowrap
        ws.cell(r, 3).border = thin_left_medium

        c = ws.cell(r, 5, "Sexo/Sex")
        c.font = Font(name=arial, bold=True, size=10)
        c.alignment = center_nowrap
        ws.cell(r, 5).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        c = ws.cell(r, 6, "Age/Edad")
        c.font = Font(name=arial, bold=True, size=10)
        c.alignment = center_nowrap
        ws.cell(r, 6).border = Border(right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        ws.merge_cells(start_row=r, start_column=7, end_row=r, end_column=8)
        c = ws.cell(r, 7, "Fecha de Elaboración (d/m/a)")
        c.font = Font(name=arial, bold=True, size=10)
        c.alignment = center_wrap
        _apply_border(ws, r, 7, Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')))
        _apply_border(ws, r, 8, Border(right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')))

        _apply_border(ws, r, 9, thin)
        _apply_border(ws, r, 10, thin)
        r += 1

        # === ROW 4-5: Values + Procedencia header ===
        row4 = r
        ws.merge_cells(start_row=r, start_column=1, end_row=r+1, end_column=2)
        c = ws.cell(r, 1, d.get("nombre", ""))
        c.font = Font(name=arial, size=14)
        c.alignment = center_wrap
        _apply_border(ws, r, 1, Border(left=Side(style='thin'), right=Side(style='thin'), bottom=Side(style='thin')))
        _apply_border(ws, r+1, 1, Border(left=Side(style='thin'), bottom=Side(style='thin')))
        _apply_border(ws, r, 2, Border(right=Side(style='thin')))
        _apply_border(ws, r+1, 2, Border(right=Side(style='thin'), bottom=Side(style='thin')))

        ws.merge_cells(start_row=r, start_column=3, end_row=r+1, end_column=4)
        c = ws.cell(r, 3, d.get("apellido", ""))
        c.font = Font(name=arial, size=15)
        c.alignment = center_wrap
        _apply_border(ws, r, 3, Border(left=Side(style='thin'), right=Side(style='thin'), bottom=Side(style='thin')))
        _apply_border(ws, r+1, 3, Border(left=Side(style='thin'), bottom=Side(style='thin')))
        _apply_border(ws, r, 4, Border(right=Side(style='thin')))
        _apply_border(ws, r+1, 4, Border(right=Side(style='thin'), bottom=Side(style='thin')))

        ws.merge_cells(start_row=r, start_column=5, end_row=r+1, end_column=5)
        c = ws.cell(r, 5, d.get("sexo", ""))
        c.font = Font(name=arial, size=16)
        c.alignment = center_wrap
        _apply_border(ws, r, 5, Border(left=Side(style='thin'), bottom=Side(style='thin')))
        _apply_border(ws, r+1, 5, Border(left=Side(style='thin'), bottom=Side(style='thin')))

        ws.merge_cells(start_row=r, start_column=6, end_row=r+1, end_column=6)
        c = ws.cell(r, 6, str(d.get("edad", "")))
        c.font = Font(name=arial, bold=True, size=16)
        c.alignment = center_wrap
        _apply_border(ws, r, 6, Border(left=Side(style='thin'), bottom=Side(style='thin')))
        _apply_border(ws, r+1, 6, Border(left=Side(style='thin'), bottom=Side(style='thin')))

        ws.merge_cells(start_row=r, start_column=7, end_row=r+1, end_column=8)
        c = ws.cell(r, 7, str(d.get("fecha_elaboracion", "")))
        c.font = Font(name=arial, size=16)
        c.alignment = center_wrap
        _apply_border(ws, r, 7, Border(left=Side(style='thin'), right=Side(style='thin'), bottom=Side(style='thin')))
        _apply_border(ws, r+1, 7, Border(left=Side(style='thin'), bottom=Side(style='thin')))
        _apply_border(ws, r, 8, Border(right=Side(style='thin')))
        _apply_border(ws, r+1, 8, Border(right=Side(style='thin'), bottom=Side(style='thin')))

        ws.merge_cells(start_row=r, start_column=9, end_row=r+1, end_column=9)
        c = ws.cell(r, 9, "Procedencia")
        c.font = Font(name=arial, bold=True, size=10)
        c.fill = yellow_fill
        c.alignment = center_nowrap
        _apply_border(ws, r, 9, thin)
        _apply_border(ws, r+1, 9, Border(left=Side(style='thin'), right=Side(style='thin'), bottom=Side(style='thin')))
        r += 2

        # === ROW 6: Labels ===
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        c = ws.cell(r, 1, "Nº Identidad")
        c.font = Font(name=arial, bold=True, size=10)
        c.alignment = center_wrap
        _apply_border(ws, r, 1, Border(left=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')))
        _apply_border(ws, r, 2, Border(top=Side(style='thin'), bottom=Side(style='thin')))

        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        c = ws.cell(r, 3, "Persona Responsable")
        c.font = Font(name=arial, bold=True, size=10)
        c.alignment = center_wrap
        _apply_border(ws, r, 3, Border(left=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')))
        _apply_border(ws, r, 4, Border(top=Side(style='thin'), bottom=Side(style='thin')))

        c = ws.cell(r, 5, "Albergue")
        c.font = Font(name=arial, bold=True, size=10)
        c.alignment = center_wrap
        _apply_border(ws, r, 5, Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')))

        c = ws.cell(r, 6, "Perfil")
        c.font = Font(name=arial, bold=True, size=10)
        c.alignment = center_wrap
        _apply_border(ws, r, 6, Border(left=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')))

        c = ws.cell(r, 7, "Teléfono")
        c.font = Font(name=arial, bold=True, size=10)
        c.alignment = center_wrap
        _apply_border(ws, r, 7, Border(left=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')))

        c = ws.cell(r, 8, "Expediente")
        c.font = Font(name=arial, bold=True, size=10)
        c.alignment = center_wrap
        _apply_border(ws, r, 8, Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')))

        _apply_border(ws, r, 9, Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')))
        ws.cell(r, 9).alignment = center_nowrap
        r += 1

        # === ROW 7-8: Values ===
        row7 = r
        ws.merge_cells(start_row=r, start_column=1, end_row=r+1, end_column=2)
        c = ws.cell(r, 1, d.get("identidad", ""))
        c.font = Font(name=arial, size=14)
        c.alignment = center_wrap
        _apply_borders_range(ws, r, 1, r, 2, thin)
        _apply_border(ws, r+1, 1, Border(left=Side(style='thin'), bottom=Side(style='thin')))
        _apply_border(ws, r+1, 2, Border(right=Side(style='thin'), bottom=Side(style='thin')))

        ws.merge_cells(start_row=r, start_column=3, end_row=r+1, end_column=4)
        c = ws.cell(r, 3, d.get("persona_responsable", ""))
        c.font = Font(name=arial, size=14)
        c.alignment = center_wrap
        _apply_borders_range(ws, r, 3, r, 4, thin)
        _apply_border(ws, r+1, 3, Border(left=Side(style='thin'), bottom=Side(style='thin')))
        _apply_border(ws, r+1, 4, Border(right=Side(style='thin'), bottom=Side(style='thin')))

        ws.merge_cells(start_row=r, start_column=5, end_row=r+1, end_column=5)
        c = ws.cell(r, 5, d.get("albergue", ""))
        c.font = Font(name=arial, size=14)
        c.alignment = center_wrap
        _apply_border(ws, r, 5, thin)
        _apply_border(ws, r+1, 5, Border(left=Side(style='thin'), right=Side(style='thin'), bottom=Side(style='thin')))

        ws.merge_cells(start_row=r, start_column=6, end_row=r+1, end_column=6)
        c = ws.cell(r, 6, d.get("perfil", ""))
        c.font = Font(name=arial, size=14)
        c.alignment = center_wrap
        _apply_border(ws, r, 6, thin)
        _apply_border(ws, r+1, 6, Border(left=Side(style='thin'), right=Side(style='thin'), bottom=Side(style='thin')))

        ws.merge_cells(start_row=r, start_column=7, end_row=r+1, end_column=7)
        c = ws.cell(r, 7, d.get("telefono", ""))
        c.font = Font(name=arial, size=14)
        c.alignment = center_wrap
        _apply_border(ws, r, 7, thin)
        _apply_border(ws, r+1, 7, Border(left=Side(style='thin'), right=Side(style='thin'), bottom=Side(style='thin')))

        ws.merge_cells(start_row=r, start_column=8, end_row=r+1, end_column=8)
        c = ws.cell(r, 8, d.get("expediente", ""))
        c.font = Font(name=arial, size=14)
        c.alignment = center_wrap
        _apply_border(ws, r, 8, thin)
        _apply_border(ws, r+1, 8, Border(left=Side(style='thin'), right=Side(style='thin'), bottom=Side(style='thin')))

        ws.merge_cells(start_row=r-1, start_column=9, end_row=r+1, end_column=9)
        c = ws.cell(r-1, 9, d.get("procedencia", ""))
        c.font = Font(name=arial, size=10)
        c.alignment = center_nowrap
        _apply_border(ws, r-1, 9, thin)
        _apply_border(ws, r, 9, Border(left=Side(style='thin'), right=Side(style='thin')))
        _apply_border(ws, r+1, 9, Border(left=Side(style='thin'), right=Side(style='thin'), bottom=Side(style='thin')))
        r += 2

        # === ROW 9-10: Domicilio ===
        ws.merge_cells(start_row=r, start_column=1, end_row=r+1, end_column=2)
        c = ws.cell(r, 1, "Domicilio del  Paciente:")
        c.font = Font(name=arial, bold=True, size=10)
        c.alignment = center_nowrap
        _apply_border(ws, r, 1, thin)
        _apply_border(ws, r+1, 1, Border(left=Side(style='thin'), bottom=Side(style='thin')))
        _apply_border(ws, r, 2, Border(right=Side(style='thin'), top=Side(style='thin')))
        _apply_border(ws, r+1, 2, Border(right=Side(style='thin'), bottom=Side(style='thin')))

        ws.merge_cells(start_row=r, start_column=3, end_row=r+1, end_column=8)
        c = ws.cell(r, 3, d.get("domicilio", ""))
        c.font = Font(name=arial, size=12)
        c.alignment = center_nowrap
        _apply_border(ws, r, 3, thin)
        _apply_borders_range(ws, r, 4, r, 8, thin_top)
        _apply_borders_range(ws, r+1, 3, r+1, 8, thin_bottom)
        r += 2

        # === ROW 11: Spacer with top border ===
        _apply_borders_range(ws, r, 1, r, 8, thin_top)
        r += 1

        # === ROW 12: HEA header ===
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        c = ws.cell(r, 1, "History of Present Illness/Historia de Enfermedad Actual:      ")
        c.font = Font(name=arial, bold=True, size=10)
        c.alignment = center_wrap
        _apply_borders_range(ws, r, 1, r, 8, thin_top_bottom)
        _apply_border(ws, r, 1, thin)
        _apply_border(ws, r, 8, thin)
        r += 1

        # === ROW 13-17: HEA value ===
        ws.merge_cells(start_row=r, start_column=1, end_row=r+4, end_column=8)
        c = ws.cell(r, 1, d.get("historia_enfermedad", ""))
        c.font = Font(name=arial, size=14)
        c.alignment = center_wrap
        _apply_border(ws, r, 1, thin)
        _apply_borders_range(ws, r, 2, r, 8, thin_top)
        _apply_borders_range(ws, r+4, 1, r+4, 8, thin_bottom)
        _apply_borders_range(ws, r, 1, r+4, 1, thin_left)
        _apply_border(ws, r+4, 8, Border(right=Side(style='thin'), bottom=Side(style='thin')))
        r += 5

        # === ROW 18: Spacer ===
        r += 1

        # === ROW 19: Medical History header ===
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        c = ws.cell(r, 1, "Medical History/Antecedentes:")
        c.font = Font(name=arial, bold=True, size=10)
        c.alignment = center_wrap
        _apply_borders_range(ws, r, 1, r, 8, thin_top_bottom)
        _apply_border(ws, r, 1, thin)
        _apply_border(ws, r, 8, thin)
        r += 1

        # === ROW 20-21: Previous illness ===
        ws.merge_cells(start_row=r, start_column=1, end_row=r+1, end_column=3)
        c = ws.cell(r, 1, "Previous illness/ Enfermedades anteriores:")
        c.font = Font(name=arial, bold=True, size=10)
        c.alignment = center_nowrap
        _apply_border(ws, r, 1, thin)
        _apply_border(ws, r, 2, thin_top)
        _apply_border(ws, r, 3, thin)
        _apply_border(ws, r+1, 1, Border(left=Side(style='thin'), bottom=Side(style='thin')))
        _apply_border(ws, r+1, 2, thin_bottom)
        _apply_border(ws, r+1, 3, Border(right=Side(style='thin'), bottom=Side(style='thin')))

        ws.merge_cells(start_row=r, start_column=4, end_row=r+1, end_column=8)
        c = ws.cell(r, 4, d.get("enfermedades_previas", ""))
        c.font = Font(name=arial, size=14)
        c.alignment = center_wrap
        _apply_border(ws, r, 4, thin)
        _apply_borders_range(ws, r, 5, r, 8, thin_top)
        _apply_borders_range(ws, r+1, 4, r+1, 8, thin_bottom)
        r += 2

        # === ROW 22-23: Past surgeries ===
        ws.merge_cells(start_row=r, start_column=1, end_row=r+1, end_column=3)
        c = ws.cell(r, 1, "Past surgeries/ Cirugías anteriores:")
        c.font = Font(name=arial, bold=True, size=10)
        c.alignment = center_nowrap
        _apply_border(ws, r, 1, thin)
        _apply_border(ws, r, 2, thin_top)
        _apply_border(ws, r, 3, thin)
        _apply_border(ws, r+1, 1, Border(left=Side(style='thin'), bottom=Side(style='thin')))
        _apply_border(ws, r+1, 2, thin_bottom)
        _apply_border(ws, r+1, 3, Border(right=Side(style='thin'), bottom=Side(style='thin')))

        ws.merge_cells(start_row=r, start_column=4, end_row=r+1, end_column=8)
        c = ws.cell(r, 4, d.get("cirugias_previas", ""))
        c.font = Font(name=arial, size=14)
        c.alignment = center_wrap
        _apply_border(ws, r, 4, thin)
        _apply_borders_range(ws, r, 5, r, 8, thin_top)
        _apply_borders_range(ws, r+1, 4, r+1, 8, thin_bottom)
        r += 2

        # === ROW 24-25: Allergies ===
        ws.merge_cells(start_row=r, start_column=1, end_row=r+1, end_column=3)
        c = ws.cell(r, 1, "Allergies/Alergias:")
        c.font = Font(name=arial, bold=True, size=10)
        c.alignment = center_nowrap
        _apply_border(ws, r, 1, thin)
        _apply_border(ws, r, 2, thin_top)
        _apply_border(ws, r, 3, thin)
        _apply_border(ws, r+1, 1, Border(left=Side(style='thin'), bottom=Side(style='thin')))
        _apply_border(ws, r+1, 2, thin_bottom)
        _apply_border(ws, r+1, 3, Border(right=Side(style='thin'), bottom=Side(style='thin')))

        ws.merge_cells(start_row=r, start_column=4, end_row=r+1, end_column=8)
        c = ws.cell(r, 4, d.get("alergias", ""))
        c.font = Font(name=arial, size=14)
        c.alignment = center_wrap
        _apply_border(ws, r, 4, thin)
        _apply_borders_range(ws, r, 5, r, 8, thin_top)
        _apply_borders_range(ws, r+1, 4, r+1, 8, thin_bottom)
        r += 2

        # === ROW 26-27: Other ===
        ws.merge_cells(start_row=r, start_column=1, end_row=r+1, end_column=3)
        c = ws.cell(r, 1, "Other/Otros Antecedentes")
        c.font = Font(name=arial, bold=True, size=10)
        c.alignment = center_wrap
        _apply_border(ws, r, 1, thin)
        _apply_border(ws, r, 2, thin_top)
        _apply_border(ws, r, 3, thin)
        _apply_border(ws, r+1, 1, Border(left=Side(style='thin'), bottom=Side(style='thin')))
        _apply_border(ws, r+1, 2, thin_bottom)
        _apply_border(ws, r+1, 3, Border(right=Side(style='thin'), bottom=Side(style='thin')))

        ws.merge_cells(start_row=r, start_column=4, end_row=r+1, end_column=8)
        c = ws.cell(r, 4, d.get("otros_antecedentes", ""))
        c.font = Font(name=arial, size=14)
        c.alignment = center_wrap
        _apply_border(ws, r, 4, thin)
        _apply_borders_range(ws, r, 5, r, 8, thin_top)
        _apply_borders_range(ws, r+1, 4, r+1, 8, thin_bottom)
        r += 2

        # === ROW 28: blank ===
        r += 1

        # === ROW 29-30: Vital signs ===
        ws.merge_cells(start_row=r, start_column=1, end_row=r+1, end_column=1)
        c = ws.cell(r, 1, "P.A./.B P:")
        c.font = Font(name=arial, bold=True, size=10)
        c.alignment = center_wrap
        _apply_border(ws, r, 1, thin)
        _apply_border(ws, r+1, 1, Border(left=Side(style='thin'), right=Side(style='thin'), bottom=Side(style='thin')))

        ws.merge_cells(start_row=r, start_column=2, end_row=r+1, end_column=2)
        c = ws.cell(r, 2, d.get("presion_arterial", ""))
        c.font = Font(name=arial, size=10)
        c.alignment = center_wrap
        _apply_border(ws, r, 2, thin)
        _apply_border(ws, r+1, 2, Border(left=Side(style='thin'), right=Side(style='thin'), bottom=Side(style='thin')))

        ws.merge_cells(start_row=r, start_column=3, end_row=r+1, end_column=3)
        fc_val = d.get("fc", "")
        c = ws.cell(r, 3, f"F.C.: {fc_val}" if fc_val else "")
        c.font = Font(name=arial, bold=True, size=10)
        c.alignment = center_wrap
        _apply_border(ws, r, 3, thin)
        _apply_border(ws, r+1, 3, Border(left=Side(style='thin'), right=Side(style='thin'), bottom=Side(style='thin')))

        pulso_val = d.get("pulso", "")
        c = ws.cell(r, 4, f"Pulso: {pulso_val}" if pulso_val else "")
        c.font = Font(name=arial, bold=True, size=9)
        c.alignment = center_wrap
        _apply_border(ws, r, 4, thin)

        ws.merge_cells(start_row=r, start_column=5, end_row=r+1, end_column=5)
        temp_val = d.get("temperatura", "")
        if temp_val and not temp_val.endswith('°C'):
            temp_str = f"T°: {temp_val}°C"
        else:
            temp_str = f"T°: {temp_val}" if temp_val else ""
        c = ws.cell(r, 5, temp_str)
        c.font = Font(name=arial, bold=True, size=11.5)
        c.alignment = center_wrap
        _apply_border(ws, r, 5, thin)
        _apply_border(ws, r+1, 5, Border(left=Side(style='thin'), right=Side(style='thin'), bottom=Side(style='thin')))

        ws.merge_cells(start_row=r, start_column=6, end_row=r+1, end_column=6)
        fr_val = d.get("fr", "")
        c = ws.cell(r, 6, f"F.R.: {fr_val}" if fr_val else "")
        c.font = Font(name=arial, bold=True, size=10)
        c.alignment = center_wrap
        _apply_border(ws, r, 6, Border(right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')))
        _apply_border(ws, r+1, 6, Border(right=Side(style='thin'), bottom=Side(style='thin')))

        c = ws.cell(r, 7, "Peso/Weight:")
        c.font = Font(name=arial, bold=True, size=10)
        c.alignment = left_center_wrap
        _apply_border(ws, r, 7, thin)

        c = ws.cell(r, 8, d.get("peso", ""))
        c.font = Font(name=arial, bold=True, size=10)
        c.alignment = left_center_wrap
        _apply_border(ws, r, 8, thin)

        r += 1

        talla_val = d.get("talla", "")
        c = ws.cell(r, 4, f"Talla: {talla_val}" if talla_val else "")
        c.font = Font(name=arial, bold=True, size=9)
        c.alignment = center_wrap
        _apply_border(ws, r, 4, thin)

        c = ws.cell(r, 7, "B.M.I.:")
        c.font = Font(name=arial, bold=True, size=10)
        c.alignment = right_center_wrap
        _apply_border(ws, r, 7, Border(left=Side(style='thin'), right=Side(style='thin'), bottom=Side(style='thin')))

        c = ws.cell(r, 8, d.get("bmi", ""))
        c.font = Font(name=arial, bold=True, size=9)
        c.alignment = left_center_wrap
        _apply_border(ws, r, 8, Border(left=Side(style='thin'), right=Side(style='thin'), bottom=Side(style='thin')))
        r += 1

        # === ROW 31: Physical Exam header ===
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        c = ws.cell(r, 1, "Physical Exam /Examen Físico ")
        c.font = Font(name=arial, bold=True, size=10)
        c.alignment = center_nowrap
        _apply_borders_range(ws, r, 1, r, 8, thin_top_bottom)
        _apply_border(ws, r, 1, thin)
        _apply_border(ws, r, 8, thin)
        r += 1

        # === ROW 32-36: Physical exam value ===
        ws.merge_cells(start_row=r, start_column=1, end_row=r+4, end_column=8)
        c = ws.cell(r, 1, d.get("examen_fisico", ""))
        c.font = Font(name=arial, size=12)
        c.alignment = center_wrap
        _apply_border(ws, r, 1, thin)
        _apply_borders_range(ws, r, 2, r, 8, thin_top)
        _apply_borders_range(ws, r+4, 1, r+4, 8, thin_bottom)
        _apply_borders_range(ws, r, 1, r+4, 1, thin_left)
        _apply_border(ws, r+4, 8, Border(right=Side(style='thin'), bottom=Side(style='thin')))
        r += 5

        # === ROW 37: Diagnosis header ===
        c = ws.cell(r, 4, "Diagnosis/ Diagnóstico")
        c.font = Font(name=arial, bold=True, size=10)
        r += 1

        # === ROW 38-42: Diagnosis value ===
        ws.merge_cells(start_row=r, start_column=1, end_row=r+4, end_column=8)
        c = ws.cell(r, 1, d.get("diagnostico", ""))
        c.font = Font(name=arial, bold=True, size=14)
        c.alignment = center_wrap
        _apply_border(ws, r, 1, thin)
        _apply_borders_range(ws, r, 2, r, 8, thin_top)
        _apply_borders_range(ws, r+4, 1, r+4, 8, thin_bottom)
        _apply_borders_range(ws, r, 1, r+4, 1, thin_left)
        _apply_border(ws, r+4, 8, Border(right=Side(style='thin'), bottom=Side(style='thin')))
        r += 5

        # === ROW 43: blank ===
        r += 1

        # === ROW 44: Doctor name ===
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
        c = ws.cell(r, 2, d.get("nombre_medico", ""))
        c.font = Font(name=arial, size=14)
        c.alignment = center_wrap
        _apply_borders_range(ws, r, 2, r, 7, medium_bottom)
        r += 1

        # === ROW 45: Doctor label ===
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
        c = ws.cell(r, 2, "    Nombre del Médico")
        c.font = Font(name=arial, bold=True, size=10)
        c.alignment = center_nowrap
        _apply_borders_range(ws, r, 2, r, 7, medium_top)
        r += 3

        # === ROW 48: Surgeon ===
        c = ws.cell(r, 1, "Surgeon/ Cirujano:")
        c.font = Font(name=arial, bold=True, size=10)
        _apply_border(ws, r, 1, medium_bottom)
        _apply_border(ws, r, 2, medium_bottom)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
        c = ws.cell(r, 3, d.get("cirujano", ""))
        c.font = Font(name=arial, size=10)
        c.alignment = center_wrap
        _apply_borders_range(ws, r, 3, r, 8, medium_bottom)
        r += 2

        # === ROW 50: Surgery Date ===
        surgery_date = d.get("fecha_cirugia", "")
        c = ws.cell(r, 1, f"Surgery Date/ Day of the Week   Fecha de Cirugía/Día de la Semana:   {surgery_date}")
        c.font = Font(name=arial, bold=True, size=10)
        _apply_borders_range(ws, r, 1, r, 5, medium_bottom)
        ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=8)
        c = ws.cell(r, 6, "")
        c.font = Font(name=arial, size=10)
        c.alignment = center_wrap
        _apply_borders_range(ws, r, 6, r, 8, medium_bottom)

    # Column widths
    for col_letter, w in col_widths.items():
        ws.column_dimensions[col_letter].width = w

    # Row heights (relative to each record block)
    for rel_row, h in row_heights.items():
        for row_idx in range(len(records)):
            actual_row = (row_idx * 51) + rel_row
            if actual_row <= ws.max_row:
                ws.row_dimensions[actual_row].height = h

    wb.save(filepath)


def _apply_border(ws, row, col, border):
    ws.cell(row, col).border = border


def _apply_borders_range(ws, r1, c1, r2, c2, border):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(r, c).border = border
